import pandas as pd
import numpy as np
from typing import List, Dict, Any, Tuple, Optional, Union
import os
import re
import logging
from datetime import datetime
import hashlib
from pathlib import Path
import tempfile
import traceback
from dataclasses import dataclass
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

# Suppress pandas warnings for cleaner output
warnings.filterwarnings('ignore', category=UserWarning)

@dataclass
class FileMetadata:
    """Metadata for processed Excel files"""
    filename: str
    file_size: int
    sheet_count: int
    total_rows: int
    processing_time: float
    checksum: str
    errors: List[str]
    warnings: List[str]

@dataclass
class MergeResult:
    """Result of the merge operation"""
    success: bool
    output_path: Optional[str]
    metadata: List[FileMetadata]
    total_rows: int
    total_columns: int
    processing_time: float
    errors: List[str]
    warnings: List[str]

class ExcelMergerError(Exception):
    """Custom exception for Excel merger operations"""
    pass

class ExcelFileMerger:
    """
    Comprehensive Excel file merger with validation, error handling, and formatting preservation
    """
    
    # Constants
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    MAX_FILES = 20
    VALID_EXTENSIONS = ['.xlsx', '.xls']
    CHUNK_SIZE = 10000  # Process data in chunks to manage memory
    
    def __init__(self, output_dir: str = None):
        """
        Initialize the Excel merger
        
        Args:
            output_dir: Directory to save output files (default: temp directory)
        """
        self.output_dir = output_dir or tempfile.gettempdir()
        self.logger = self._setup_logger()
        self.processed_files: List[FileMetadata] = []
        
    def _setup_logger(self) -> logging.Logger:
        """Setup logging configuration"""
        logger = logging.getLogger('ExcelMerger')
        logger.setLevel(logging.INFO)
        
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            
        return logger
    
    def _calculate_checksum(self, file_path: str) -> str:
        """Calculate MD5 checksum of file"""
        hash_md5 = hashlib.md5()
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception as e:
            self.logger.error(f"Error calculating checksum for {file_path}: {str(e)}")
            return ""
    
    def _validate_file_format(self, file_path: str) -> Tuple[bool, List[str]]:
        """
        Validate if file is a valid Excel format
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (is_valid, error_messages)
        """
        errors = []
        
        # Check file extension
        file_ext = Path(file_path).suffix.lower()
        if file_ext not in self.VALID_EXTENSIONS:
            errors.append(f"Invalid file extension: {file_ext}. Expected: {', '.join(self.VALID_EXTENSIONS)}")
            return False, errors
        
        # Check file size
        file_size = os.path.getsize(file_path)
        if file_size > self.MAX_FILE_SIZE:
            errors.append(f"File size ({file_size / 1024 / 1024:.2f}MB) exceeds maximum allowed size ({self.MAX_FILE_SIZE / 1024 / 1024}MB)")
            return False, errors
        
        # Check if file is a valid Excel file
        try:
            with pd.ExcelFile(file_path, engine='openpyxl' if file_ext == '.xlsx' else 'xlrd') as excel_file:
                if not excel_file.sheet_names:
                    errors.append("File contains no sheets")
                    return False, errors
        except Exception as e:
            errors.append(f"Invalid Excel file or corrupted: {str(e)}")
            return False, errors
        
        return True, errors
    
    def _standardize_column_name(self, column_name: str) -> str:
        """
        Standardize column names by removing special characters and converting to lowercase
        
        Args:
            column_name: Original column name
            
        Returns:
            Standardized column name
        """
        if pd.isna(column_name) or column_name is None:
            return "unnamed_column"
        
        # Convert to string and strip whitespace
        name = str(column_name).strip()
        
        # Remove special characters and replace with underscore
        name = re.sub(r'[^a-zA-Z0-9\s]', '_', name)
        
        # Replace multiple spaces/underscores with single underscore
        name = re.sub(r'[\s_]+', '_', name)
        
        # Convert to lowercase and remove leading/trailing underscores
        name = name.lower().strip('_')
        
        # Ensure name is not empty
        if not name:
            name = "unnamed_column"
            
        return name
    
    def _process_dataframe(self, df: pd.DataFrame, source_file: str, sheet_name: str) -> pd.DataFrame:
        """
        Process and clean a DataFrame
        
        Args:
            df: Input DataFrame
            source_file: Source file name
            sheet_name: Sheet name
            
        Returns:
            Processed DataFrame
        """
        # Add metadata columns
        df = df.copy()
        df['_source_file'] = source_file
        df['_source_sheet'] = sheet_name
        df['_processed_timestamp'] = datetime.now()
        
        # Standardize column names
        new_columns = {}
        for col in df.columns:
            if col.startswith('_'):  # Skip metadata columns
                continue
            standardized = self._standardize_column_name(col)
            new_columns[col] = standardized
        
        df = df.rename(columns=new_columns)
        
        # Handle duplicate column names
        columns = list(df.columns)
        seen = set()
        for i, col in enumerate(columns):
            if col in seen and not col.startswith('_'):
                counter = 1
                new_col = f"{col}_{counter}"
                while new_col in seen:
                    counter += 1
                    new_col = f"{col}_{counter}"
                columns[i] = new_col
            seen.add(columns[i])
        
        df.columns = columns
        
        return df
    
    def _extract_sheets_from_file(self, file_path: str) -> Tuple[List[pd.DataFrame], FileMetadata]:
        """
        Extract all sheets from an Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Tuple of (list of DataFrames, file metadata)
        """
        start_time = datetime.now()
        filename = Path(file_path).name
        file_size = os.path.getsize(file_path)
        checksum = self._calculate_checksum(file_path)
        errors = []
        warnings = []
        dataframes = []
        total_rows = 0
        
        try:
            # Determine engine based on file extension
            engine = 'openpyxl' if file_path.endswith('.xlsx') else 'xlrd'
            
            with pd.ExcelFile(file_path, engine=engine) as excel_file:
                sheet_count = len(excel_file.sheet_names)
                
                for sheet_name in excel_file.sheet_names:
                    try:
                        # Read sheet with error handling
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0)
                        
                        # Skip empty sheets
                        if df.empty:
                            warnings.append(f"Sheet '{sheet_name}' is empty and will be skipped")
                            continue
                        
                        # Remove completely empty rows and columns
                        df = df.dropna(how='all').dropna(axis=1, how='all')
                        
                        if df.empty:
                            warnings.append(f"Sheet '{sheet_name}' contains no data after cleaning")
                            continue
                        
                        # Process the DataFrame
                        processed_df = self._process_dataframe(df, filename, sheet_name)
                        dataframes.append(processed_df)
                        total_rows += len(processed_df)
                        
                        self.logger.info(f"Processed sheet '{sheet_name}' from '{filename}': {len(processed_df)} rows")
                        
                    except Exception as e:
                        error_msg = f"Error processing sheet '{sheet_name}': {str(e)}"
                        errors.append(error_msg)
                        self.logger.error(error_msg)
                        continue
                
        except Exception as e:
            error_msg = f"Error reading file '{filename}': {str(e)}"
            errors.append(error_msg)
            self.logger.error(error_msg)
            raise ExcelMergerError(error_msg)
        
        processing_time = (datetime.now() - start_time).total_seconds()
        
        metadata = FileMetadata(
            filename=filename,
            file_size=file_size,
            sheet_count=len(excel_file.sheet_names) if 'excel_file' in locals() else 0,
            total_rows=total_rows,
            processing_time=processing_time,
            checksum=checksum,
            errors=errors,
            warnings=warnings
        )
        
        return dataframes, metadata
    
    def _merge_dataframes(self, all_dataframes: List[pd.DataFrame]) -> pd.DataFrame:
        """
        Merge all DataFrames into a single consolidated DataFrame
        
        Args:
            all_dataframes: List of DataFrames to merge
            
        Returns:
            Consolidated DataFrame
        """
        if not all_dataframes:
            raise ExcelMergerError("No valid data found to merge")
        
        try:
            # Concatenate all DataFrames
            merged_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
            
            # Remove duplicate rows (excluding metadata columns)
            data_columns = [col for col in merged_df.columns if not col.startswith('_')]
            merged_df = merged_df.drop_duplicates(subset=data_columns, keep='first')
            
            # Sort by source file and sheet for better organization
            merged_df = merged_df.sort_values(['_source_file', '_source_sheet'], na_position='last')
            merged_df = merged_df.reset_index(drop=True)
            
            self.logger.info(f"Successfully merged {len(all_dataframes)} DataFrames into {len(merged_df)} rows")
            
            return merged_df
            
        except Exception as e:
            error_msg = f"Error merging DataFrames: {str(e)}"
            self.logger.error(error_msg)
            raise ExcelMergerError(error_msg)
    
    def _create_metadata_sheet(self, metadata_list: List[FileMetadata], merged_df: pd.DataFrame) -> pd.DataFrame:
        """
        Create a metadata sheet with processing information
        
        Args:
            metadata_list: List of file metadata
            merged_df: Merged DataFrame
            
        Returns:
            Metadata DataFrame
        """
        metadata_data = []
        
        for metadata in metadata_list:
            metadata_data.append({
                'Source_File': metadata.filename,
                'File_Size_MB': round(metadata.file_size / 1024 / 1024, 2),
                'Sheet_Count': metadata.sheet_count,
                'Rows_Processed': metadata.total_rows,
                'Processing_Time_Seconds': round(metadata.processing_time, 2),
                'Checksum': metadata.checksum,
                'Errors': '; '.join(metadata.errors) if metadata.errors else 'None',
                'Warnings': '; '.join(metadata.warnings) if metadata.warnings else 'None'
            })
        
        # Add summary information
        summary_data = {
            'Source_File': 'SUMMARY',
            'File_Size_MB': sum(m.file_size for m in metadata_list) / 1024 / 1024,
            'Sheet_Count': sum(m.sheet_count for m in metadata_list),
            'Rows_Processed': len(merged_df),
            'Processing_Time_Seconds': sum(m.processing_time for m in metadata_list),
            'Checksum': 'N/A',
            'Errors': str(sum(len(m.errors) for m in metadata_list)),
            'Warnings': str(sum(len(m.warnings) for m in metadata_list))
        }
        
        metadata_data.append(summary_data)
        
        return pd.DataFrame(metadata_data)
    
    def _save_to_excel(self, merged_df: pd.DataFrame, metadata_df: pd.DataFrame, output_path: str) -> None:
        """
        Save merged data and metadata to Excel file with formatting
        
        Args:
            merged_df: Merged data DataFrame
            metadata_df: Metadata DataFrame
            output_path: Output file path
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main data
                merged_df.to_excel(writer, sheet_name='Merged_Data', index=False)
                
                # Write metadata
                metadata_df.to_excel(writer, sheet_name='Processing_Metadata', index=False)
                
                # Apply formatting
                workbook = writer.book
                
                # Format main data sheet
                data_sheet = workbook['Merged_Data']
                self._apply_sheet_formatting(data_sheet, is_data_sheet=True)
                
                # Format metadata sheet
                metadata_sheet = workbook['Processing_Metadata']
                self._apply_sheet_formatting(metadata_sheet, is_data_sheet=False)
            
            self.logger.info(f"Successfully saved merged file to: {output_path}")
            
        except Exception as e:
            error_msg = f"Error saving Excel file: {str(e)}"
            self.logger.error(error_msg)
            raise ExcelMergerError(error_msg)
    
    def _apply_sheet_formatting(self, sheet, is_data_sheet: bool = True) -> None:
        """
        Apply formatting to Excel sheet
        
        Args:
            sheet: Openpyxl worksheet object
            is_data_sheet: Whether this is the main data sheet
        """
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting to first row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
    
    def merge_excel_files(self, file_paths: List[str]) -> MergeResult:
        """
        Main function to merge multiple Excel files
        
        Args:
            file_paths: List of paths to Excel files
            
        Returns:
            MergeResult object containing operation results
        """
        start_time = datetime.now()
        all_errors = []
        all_warnings = []
        
        try:
            # Validate input parameters
            if not file_paths:
                raise ExcelMergerError("No files provided for merging")
            
            if len(file_paths) > self.MAX_FILES:
                raise ExcelMergerError(f"Too many files provided. Maximum allowed: {self.MAX_FILES}")
            
            # Validate all files first
            valid_files = []
            for file_path in file_paths:
                if not os.path.exists(file_path):
                    error_msg = f"File not found: {file_path}"
                    all_errors.append(error_msg)
                    continue
                
                is_valid, errors = self._validate_file_format(file_path)
                if not is_valid:
                    all_errors.extend([f"{Path(file_path).name}: {error}" for error in errors])
                    continue
                
                valid_files.append(file_path)
            
            if not valid_files:
                raise ExcelMergerError("No valid Excel files found")
            
            # Process each file
            all_dataframes = []
            metadata_list = []
            
            for file_path in valid_files:
                try:
                    dataframes, metadata = self._extract_sheets_from_file(file_path)
                    all_dataframes.extend(dataframes)
                    metadata_list.append(metadata)
                    
                    # Collect errors and warnings
                    all_errors.extend(metadata.errors)
                    all_warnings.extend(metadata.warnings)
                    
                except ExcelMergerError:
                    # Re-raise ExcelMergerError
                    raise
                except Exception as e:
                    error_msg = f"Unexpected error processing {Path(file_path).name}: {str(e)}"
                    all_errors.append(error_msg)
                    self.logger.error(error_msg)
                    continue
            
            if not all_dataframes:
                raise ExcelMergerError("No valid data found in any of the provided files")
            
            # Merge all DataFrames
            merged_df = self._merge_dataframes(all_dataframes)
            
            # Create metadata sheet
            metadata_df = self._create_metadata_sheet(metadata_list, merged_df)
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"merged_excel_files_{timestamp}.xlsx"
            output_path = os.path.join(self.output_dir, output_filename)
            
            # Save to Excel
            self._save_to_excel(merged_df, metadata_df, output_path)
            
            # Calculate final statistics
            processing_time = (datetime.now() - start_time).total_seconds()
            
            result = MergeResult(
                success=True,
                output_path=output_path,
                metadata=metadata_list,
                total_rows=len(merged_df),
                total_columns=len(merged_df.columns),
                processing_time=processing_time,
                errors=all_errors,
                warnings=all_warnings
            )
            
            self.logger.info(f"Merge operation completed successfully in {processing_time:.2f} seconds")
            self.logger.info(f"Output file: {output_path}")
            self.logger.info(f"Total rows: {result.total_rows}, Total columns: {result.total_columns}")
            
            return result
            
        except ExcelMergerError as e:
            processing_time = (datetime.now() - start_time).total_seconds()
            self.logger.error(f"Merge operation failed: {str(e)}")
            
            return MergeResult(
                success=False,
                output_path=None,
                metadata=metadata_list if 'metadata_list' in locals() else [],
                total_rows=0,
                total_columns=0,
                processing_time=processing_time,
                errors=all_errors + [str(e)],
                warnings=all_warnings
            )
        
        except Exception as e:
            processing_time = (datetime.now() - start_time).total_seconds()
            error_msg = f"Unexpected error during merge operation: {str(e)}"
            self.logger.error(error_msg)
            self.logger.error(traceback.format_exc())
            
            return MergeResult(
                success=False,
                output_path=None,
                metadata=metadata_list if 'metadata_list' in locals() else [],
                total_rows=0,
                total_columns=0,
                processing_time=processing_time,
                errors=all_errors + [error_msg],
                warnings=all_warnings
            )

# Convenience function for direct usage
def merge_excel_files(file_paths: List[str], output_dir: str = None) -> Dict[str, Any]:
    """
    Convenience function to merge Excel files
    
    Args:
        file_paths: List of paths to Excel files to merge
        output_dir: Directory to save output file (optional)
        
    Returns:
        Dictionary containing merge results
    """
    merger = ExcelFileMerger(output_dir)
    result = merger.merge_excel_files(file_paths)
    
    # Convert result to dictionary for easier JSON serialization
    return {
        'success': result.success,
        'output_path': result.output_path,
        'total_rows': result.total_rows,
        'total_columns': result.total_columns,
        'processing_time': result.processing_time,
        'errors': result.errors,
        'warnings': result.warnings,
        'metadata': [
            {
                'filename': m.filename,
                'file_size_mb': round(m.file_size / 1024 / 1024, 2),
                'sheet_count': m.sheet_count,
                'total_rows': m.total_rows,
                'processing_time': m.processing_time,
                'checksum': m.checksum,
                'errors': m.errors,
                'warnings': m.warnings
            }
            for m in result.metadata
        ]
    }

# Example usage and testing
if __name__ == "__main__":
    # Example usage
    sample_files = [
        "sample_file1.xlsx",
        "sample_file2.xlsx",
        "sample_file3.xls"
    ]
    
    # Create merger instance
    merger = ExcelFileMerger(output_dir="./output")
    
    # Merge files
    result = merger.merge_excel_files(sample_files)
    
    if result.success:
        print(f"✅ Merge successful!")
        print(f"📁 Output file: {result.output_path}")
        print(f"📊 Total rows: {result.total_rows}")
        print(f"📋 Total columns: {result.total_columns}")
        print(f"⏱️ Processing time: {result.processing_time:.2f} seconds")
        
        if result.warnings:
            print(f"⚠️ Warnings: {len(result.warnings)}")
            for warning in result.warnings:
                print(f"   - {warning}")
    else:
        print(f"❌ Merge failed!")
        print(f"🚫 Errors: {len(result.errors)}")
        for error in result.errors:
            print(f"   - {error}")